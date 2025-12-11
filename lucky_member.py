# lucky_member.py
import tkinter as tk
from tkinter import filedialog, messagebox
# import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import random
# import os

def is_red(cell):
    # 判断单元格字体是否为红色（FFFF0000）
    if cell.font and cell.font.color and cell.font.color.rgb:
        # 转成字符串再 upper
        return str(cell.font.color.rgb).upper() == 'FFFF0000'
    return False

def load_and_draw(file_path):
    # 读取 Excel 并抽取幸运成员
    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("文件错误", f"无法读取文件：\n{e}")
        return

    members = []
    total_times = 0
    total_score = 0

    # 从第2行开始读取（跳过表头）
    for row in ws.iter_rows(min_row=2, values_only=False):
        name_cell = row[0]
        if name_cell.value is None:
            continue
        name = str(name_cell.value).strip()
        if not name:
            continue

        # 仅统计红色成员
        if not is_red(name_cell):
            continue

        # 计算参与次数（中间列，去掉首列名称和末列总分）
        times = 0
        for cell in row[1:-1]:
            v = cell.value
            if v is not None and str(v).isdigit():
                times += int(v)

        # 读取总分（最后一列）
        score_cell = row[-1]
        score = 0
        if score_cell.value is not None:
            try:
                score = int(score_cell.value)
            except ValueError:
                pass

        # 数据合法性检查
        if (times == 0 and score != 0) or (times != 0 and score == 0):
            messagebox.showerror(
                "数据错误",
                f"成员【{name}】的次数或分数数据错误（次数={times}，分数={score}）"
            )
            return

        members.append({
            'name': name,
            'times': times,
            'score': score
        })
        total_times += times
        total_score += score

    # 当所有红色成员未参与时报错
    if total_times == 0 and total_score == 0:
        messagebox.showerror("错误", "所有红色成员均未参与（次数与分数均为 0）！")
        return

    # 计算概率并构建抽奖池
    lottery = []
    for m in members:
        p_times = m['times'] / total_times
        p_score = m['score'] / total_score
        prob = 0.5 * p_times + 0.5 * p_score
        lottery.extend([m['name']] * max(1, int(prob * 10000)))

    lucky_one = random.choice(lottery)
    messagebox.showinfo("抽奖结果", f"恭喜幸运成员：{lucky_one}")

def select_file():
    filetypes = [("Excel 文件", "*.xlsx *.xlsm")]
    path = filedialog.askopenfilename(title="选择班组战成绩表", filetypes=filetypes)
    if path:
        load_and_draw(path)

# -------------- GUI 部分 --------------
from PIL import Image, ImageTk

WIN_W, WIN_H = 525, 300          # 窗口尺寸
IMG_PATH = '20180315-edit.png'   # 背景图
ALPHA = 0.5                     # 透明度

def set_background(win: tk.Tk, img_path, alpha=0.85):
    # 1. 让窗口无框、置顶、背景色设为透明穿透
    win.update()
    win.overrideredirect(True)
    win.attributes('-topmost', True)
    trans_color = 'grey15'
    win.configure(bg=trans_color)
    win.attributes('-transparentcolor', trans_color)

    # 2. 创建同尺寸 canvas，贴图
    canvas = tk.Canvas(win, width=WIN_W, height=WIN_H,
                       highlightthickness=0, bg=trans_color)
    canvas.pack(fill='both', expand=True)

    # 3. 读图缩放调透明度
    img = Image.open(img_path).convert('RGBA')
    img = img.resize((WIN_W, WIN_H), Image.LANCZOS)
    # 调 alpha：把每个像素的 A 通道乘系数
    *rgb, a = img.split()
    a = a.point(lambda p: int(p * alpha))
    img.putalpha(a)
    bg_img = ImageTk.PhotoImage(img)
    canvas.bg_img = bg_img          # 防止被回收
    canvas.create_image(0, 0, anchor='nw', image=bg_img)

    def start_move(event):
        win._x, win._y = event.x, event.y
    def on_move(event):
        win.geometry(f'+{event.x_root - win._x}+{event.y_root - win._y}')
    canvas.bind('<Button-1>', start_move)
    canvas.bind('<B1-Motion>', on_move)

    return canvas

# ---------------- 主程序 ----------------
root = tk.Tk()
root.geometry(f'{WIN_W}x{WIN_H}+300+200')   # 初始位置
canvas = set_background(root, IMG_PATH, ALPHA)

# 把素材全部放到 canvas 上
canvas.create_text(WIN_W//2, WIN_H//3,
                   text='选择 Excel 文件并抽奖',
                   fill='white',
                   font=('Microsoft YaHei', 12),
                   anchor='center')
btn = tk.Button(root, text='选择文件并抽奖', command=select_file, width=18, height=2)
canvas.create_window(WIN_W//2, (WIN_H//2), window=btn)
quit_btn = tk.Button(root, text='退出', command=root.quit, width=10)
canvas.create_window(WIN_W//2, (WIN_H//5)*4, window=quit_btn)

root.mainloop()